"""
Family Budget Backend API
=========================
FastAPI application for managing family incomes, expenses, and reports.
"""

import io
import json
import logging
import os
from contextlib import contextmanager, asynccontextmanager
from datetime import date, timedelta
from enum import Enum
from typing import Optional
from urllib.parse import quote
from uuid import uuid4

import bcrypt
import openpyxl
import psycopg2
import psycopg2.extras
from fastapi import FastAPI, File, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Border, Font, Side
from pydantic import BaseModel, Field

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("family_budget")

DB_CONFIG = {
    "dbname": "family_budget",
    "user": "postgres",
    "password": "123",
    "host": "localhost",
    "port": 5432,
}

ADMIN_PASSWORD_HASH = os.getenv(
    "ADMIN_PASSWORD_HASH",
    b"$2b$12$sEAJAw4ZoJShvEdfYQGGT.IETUD5x2WWqNqWjpJH0pxi7.QhwNt6.",
)

# ---------- Enums ----------
class KinshipType(str, Enum):
    mother = "мать"
    father = "отец"
    son = "сын"
    daughter = "дочь"
    grandmother = "бабушка"
    grandfather = "дедушка"

class IncomeType(str, Enum):
    salary = "зарплата"
    scholarship = "стипендия"
    pension = "пенсия"
    extra = "дополнительный заработок"

class AccountType(str, Enum):
    main = "основной счет"
    savings = "накопления"
    stash = "заначка"

class ExpenseCategory(str, Enum):
    products = "продукты"
    clothes = "одежда"
    utilities = "коммунальные услуги"
    mobile = "мобильная связь"
    leisure = "отдых"
    education = "обучение"
    entertainment = "развлечения"
    medicine = "лекарства"
    transport = "транспорт"
    other = "другое"

# ---------- Session Management ----------
class SessionManager:
    _instance = None
    def __init__(self):
        self.sessions = {}
    @classmethod
    def get_instance(cls):
        if not cls._instance:
            cls._instance = SessionManager()
        return cls._instance
    def get_user(self, token: str) -> dict:
        if token not in self.sessions:
            raise HTTPException(status_code=401, detail="Сессия недействительна")
        return self.sessions[token]
    def end_session(self, token: str) -> None:
        self.sessions.pop(token, None)
    def create_session(self, role: str, member_id: Optional[int], full_name: str) -> str:
        token = str(uuid4())
        self.sessions[token] = {"role": role, "member_id": member_id, "full_name": full_name}
        return token

session_manager = SessionManager.get_instance()

# ---------- Authorization Proxy ----------
class AuthProxy:
    def __init__(self, sm: SessionManager):
        self.sm = sm
    def check(self, token: str, required_role: str) -> dict:
        user = self.sm.get_user(token)
        if user["role"] != required_role:
            raise HTTPException(status_code=403, detail=f"Требуется роль {required_role}")
        return user

auth_proxy = AuthProxy(session_manager)

# ---------- Database Initialization ----------
def init_db() -> None:
    commands = [
        """
        CREATE TABLE IF NOT EXISTS accounts (
            id SERIAL PRIMARY KEY,
            name VARCHAR(50) NOT NULL UNIQUE
        );
        """,
        """
        CREATE TABLE IF NOT EXISTS balance_history (
            id SERIAL PRIMARY KEY,
            operation_type VARCHAR(10) NOT NULL,
            operation_id INTEGER NOT NULL,
            balance_before NUMERIC(12,2) NOT NULL,
            balance_after NUMERIC(12,2) NOT NULL,
            change_amount NUMERIC(12,2) NOT NULL,
            changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_balance_history_changed_at
        ON balance_history (changed_at);
        """,
        """
        CREATE TABLE IF NOT EXISTS expense_categories (
            id SERIAL PRIMARY KEY,
            name VARCHAR(50) NOT NULL UNIQUE
        );
        """,
        """
        CREATE TABLE IF NOT EXISTS family_members (
            id SERIAL PRIMARY KEY,
            full_name TEXT NOT NULL,
            kinship_type TEXT NOT NULL,
            CONSTRAINT family_members_kinship_type_check
            CHECK (kinship_type IN ('мать','отец','сын','дочь','бабушка','дедушка'))
        );
        """,
        """
        CREATE TABLE IF NOT EXISTS income_types (
            id SERIAL PRIMARY KEY,
            name VARCHAR(50) NOT NULL UNIQUE
        );
        """,
        """
        CREATE TABLE IF NOT EXISTS member_income_types (
            member_id INTEGER NOT NULL REFERENCES family_members(id) ON DELETE CASCADE,
            income_type VARCHAR(50) NOT NULL,
            PRIMARY KEY (member_id, income_type)
        );
        """,
        """
        CREATE TABLE IF NOT EXISTS income (
            id SERIAL PRIMARY KEY,
            member_id INTEGER NOT NULL,
            amount NUMERIC(12,2) NOT NULL CHECK (amount > 0),
            date DATE NOT NULL,
            operator_member_id INTEGER NOT NULL,
            income_type_id INTEGER NOT NULL,
            account_id INTEGER NOT NULL,
            CONSTRAINT income_member_id_fkey FOREIGN KEY (member_id)
                REFERENCES family_members(id) ON DELETE RESTRICT,
            CONSTRAINT income_operator_member_id_fkey FOREIGN KEY (operator_member_id)
                REFERENCES family_members(id) ON DELETE RESTRICT,
            CONSTRAINT income_income_type_id_fkey FOREIGN KEY (income_type_id)
                REFERENCES income_types(id),
            CONSTRAINT income_account_id_fkey FOREIGN KEY (account_id)
                REFERENCES accounts(id)
        );
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_income_date ON income(date);
        CREATE INDEX IF NOT EXISTS idx_income_date_member ON income(date, member_id);
        CREATE INDEX IF NOT EXISTS idx_income_member ON income(member_id);
        """,
        """
        CREATE TABLE IF NOT EXISTS expense (
            id SERIAL PRIMARY KEY,
            amount NUMERIC(12,2) NOT NULL CHECK (amount > 0),
            date DATE NOT NULL,
            is_planned BOOLEAN NOT NULL DEFAULT FALSE,
            operator_member_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            CONSTRAINT expense_operator_member_id_fkey FOREIGN KEY (operator_member_id)
                REFERENCES family_members(id) ON DELETE RESTRICT,
            CONSTRAINT expense_category_id_fkey FOREIGN KEY (category_id)
                REFERENCES expense_categories(id)
        );
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_expense_date ON expense(date);
        CREATE INDEX IF NOT EXISTS idx_expense_date_planned ON expense(date, is_planned);
        CREATE INDEX IF NOT EXISTS idx_expense_planned ON expense(is_planned);
        CREATE INDEX IF NOT EXISTS idx_expense_unplanned ON expense(is_planned) WHERE is_planned = false;
        """,
        """
        CREATE OR REPLACE FUNCTION fn_current_balance()
        RETURNS NUMERIC
        LANGUAGE plpgsql AS $$
        DECLARE
            total_income NUMERIC;
            total_expense NUMERIC;
        BEGIN
            SELECT COALESCE(SUM(amount), 0) INTO total_income FROM income;
            SELECT COALESCE(SUM(amount), 0) INTO total_expense FROM expense WHERE is_planned = FALSE;
            RETURN total_income - total_expense;
        END;
        $$;
        """,
        """
        CREATE OR REPLACE FUNCTION trg_log_balance_change()
        RETURNS TRIGGER
        LANGUAGE plpgsql AS $$
        DECLARE
            balance_before NUMERIC;
            balance_after NUMERIC;
            op_type TEXT;
            op_id INTEGER;
        BEGIN
            IF TG_TABLE_NAME = 'income' THEN
                op_type := 'income';
                op_id := NEW.id;
            ELSIF TG_TABLE_NAME = 'expense' AND NEW.is_planned = FALSE THEN
                op_type := 'expense';
                op_id := NEW.id;
            ELSE
                RETURN NEW;
            END IF;

            IF TG_OP = 'INSERT' THEN
                IF op_type = 'income' THEN
                    balance_before := fn_current_balance() - NEW.amount;
                ELSE
                    balance_before := fn_current_balance() + NEW.amount;
                END IF;
                balance_after := balance_before + (CASE WHEN op_type = 'income' THEN NEW.amount ELSE -NEW.amount END);
            ELSE
                RETURN NEW;
            END IF;

            INSERT INTO balance_history (operation_type, operation_id, balance_before, balance_after, change_amount)
            VALUES (op_type, op_id, balance_before, balance_after, balance_after - balance_before);

            RETURN NEW;
        END;
        $$;
        """,
        # Баланс может стать отрицательным, но после этого все расходы запрещены.
        """
        CREATE OR REPLACE FUNCTION trg_prevent_negative_balance()
        RETURNS TRIGGER
        LANGUAGE plpgsql AS $$
        DECLARE
            current_bal NUMERIC;
        BEGIN
            IF NEW.is_planned = FALSE THEN
                current_bal := fn_current_balance();
                -- Запрещаем расход, если текущий баланс уже <= 0
                IF current_bal <= 0 THEN
                    RAISE EXCEPTION 'Запрет расхода: текущий баланс уже нулевой или отрицательный (текущий баланс = %)', current_bal;
                END IF;
                -- Разрешаем расход, даже если он сделает баланс отрицательным
            END IF;
            RETURN NEW;
        END;
        $$;
        """,
        "DROP TRIGGER IF EXISTS trg_log_balance_change_income ON income;",
        "DROP TRIGGER IF EXISTS trg_log_balance_change_expense ON expense;",
        "DROP TRIGGER IF EXISTS trg_prevent_negative_balance ON expense;",
        """
        CREATE TRIGGER trg_log_balance_change_income
        AFTER INSERT ON income
        FOR EACH ROW
        EXECUTE FUNCTION trg_log_balance_change();
        """,
        """
        CREATE TRIGGER trg_log_balance_change_expense
        AFTER INSERT ON expense
        FOR EACH ROW
        EXECUTE FUNCTION trg_log_balance_change();
        """,
        """
        CREATE TRIGGER trg_prevent_negative_balance
        BEFORE INSERT ON expense
        FOR EACH ROW
        EXECUTE FUNCTION trg_prevent_negative_balance();
        """,
        """
        CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY,
            user_id INTEGER,
            user_role VARCHAR(20) NOT NULL,
            action VARCHAR(50) NOT NULL,
            details TEXT,
            entity_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    ]
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        for cmd in commands:
            cur.execute(cmd)
        conn.commit()
        cur.close()
        conn.close()
        logger.info("Database initialized successfully")
    except Exception as e:
        logger.error(f"Database initialization error: {e}")

# ---------- Database Connection ----------
@contextmanager
def db_connection():
    conn = psycopg2.connect(**DB_CONFIG)
    try:
        yield conn
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

# ---------- Helpers ----------
def _get_income_type_id(cur, name: str) -> int:
    cur.execute("SELECT id FROM income_types WHERE name = %s", (name,))
    result = cur.fetchone()
    if not result:
        cur.execute("INSERT INTO income_types (name) VALUES (%s) RETURNING id", (name,))
        return cur.fetchone()[0]
    return result[0]

def _get_account_id(cur, name: str) -> int:
    cur.execute("SELECT id FROM accounts WHERE name = %s", (name,))
    result = cur.fetchone()
    if not result:
        cur.execute("INSERT INTO accounts (name) VALUES (%s) RETURNING id", (name,))
        return cur.fetchone()[0]
    return result[0]

def _get_category_id(cur, name: str) -> int:
    cur.execute("SELECT id FROM expense_categories WHERE name = %s", (name,))
    result = cur.fetchone()
    if not result:
        cur.execute("INSERT INTO expense_categories (name) VALUES (%s) RETURNING id", (name,))
        return cur.fetchone()[0]
    return result[0]

def _add_default_income_types_for_member(cur, member_id: int) -> None:
    for inc_type in ["зарплата", "стипендия", "пенсия", "дополнительный заработок"]:
        cur.execute(
            "INSERT INTO member_income_types (member_id, income_type) VALUES (%s, %s) ON CONFLICT DO NOTHING",
            (member_id, inc_type),
        )

# ---------- Audit Logging ----------
def log_audit(user_info: dict, action: str, entity_id: Optional[int] = None, details: Optional[str] = None) -> None:
    try:
        with db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO audit_log (user_id, user_role, action, entity_id, details)
                       VALUES (%s, %s, %s, %s, %s)""",
                    (user_info.get("member_id"), user_info.get("role", "unknown"),
                     action, entity_id, details),
                )
                conn.commit()
    except Exception as e:
        logger.error(f"Failed to write audit log: {e}")

# ---------- Models ----------
class AdminLogin(BaseModel):
    username: str
    password: str

class MemberCreate(BaseModel):
    full_name: str
    kinship_type: KinshipType

class IncomeIn(BaseModel):
    income_type: IncomeType
    account: AccountType
    amount: float = Field(..., gt=0)
    date: date
    member_id: Optional[int] = None

class ExpenseIn(BaseModel):
    category: ExpenseCategory
    amount: float = Field(..., gt=0)
    date: date
    is_planned: bool = False
    member_id: Optional[int] = None

class ActivateRequest(BaseModel):
    member_id: int

# ---------- FastAPI ----------
@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield

app = FastAPI(
    title="Семейный бюджет",
    lifespan=lifespan,
    docs_url="/docs",
    redoc_url="/redoc",
    openapi_url="/openapi.json"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------- Корневой эндпоинт ----------
@app.get("/")
def root():
    return {"message": "Семейный бюджет API работает", "status": "ok"}

# ---------- Auth ----------
@app.post("/login/admin")
def login_admin(creds: AdminLogin):
    if creds.username != "admin" or not bcrypt.checkpw(creds.password.encode("utf-8"), ADMIN_PASSWORD_HASH):
        raise HTTPException(status_code=403, detail="Неверные учётные данные")
    token = session_manager.create_session("admin", None, "Администратор")
    log_audit({"role": "admin"}, "login_admin")
    return {"token": token, "full_name": "Администратор"}

@app.post("/login/guest")
def login_guest():
    token = session_manager.create_session("guest", None, "Гость")
    log_audit({"role": "guest"}, "login_guest")
    return {"token": token, "full_name": "Гость"}

@app.get("/public/members")
def public_members():
    with db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, full_name, kinship_type FROM family_members ORDER BY id")
            rows = cur.fetchall()
    return rows

@app.post("/session/activate")
def activate_session(req: ActivateRequest):
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, full_name FROM family_members WHERE id = %s", (req.member_id,))
            member = cur.fetchone()
            if not member:
                raise HTTPException(status_code=404, detail="Член семьи не найден")
    token = session_manager.create_session("client", member[0], member[1])
    log_audit({"role": "client", "member_id": member[0]}, "activate_member", entity_id=member[0])
    return {"token": token, "full_name": member[1]}

@app.post("/session/end")
def end_session(token: str = Header(...)):
    user_info = session_manager.get_user(token)
    log_audit(user_info, "logout")
    session_manager.end_session(token)
    return {"detail": "Сеанс завершён"}

# ---------- Members (admin only) ----------
@app.post("/members")
def create_member(data: MemberCreate, token: str = Header(...)):
    auth_proxy.check(token, "admin")
    with db_connection() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    "INSERT INTO family_members (full_name, kinship_type) VALUES (%s, %s) RETURNING id",
                    (data.full_name, data.kinship_type.value),
                )
                new_id = cur.fetchone()[0]
                _add_default_income_types_for_member(cur, new_id)
                conn.commit()
                log_audit({"role": "admin"}, "add_member", entity_id=new_id, details=f"Name: {data.full_name}")
            except Exception as e:
                logger.error(f"Error creating member: {e}")
                raise HTTPException(status_code=400, detail=str(e))
    return {"id": new_id}

@app.post("/members/upload")
async def upload_members(file: UploadFile = File(...), token: str = Header(...)):
    auth_proxy.check(token, "admin")
    try:
        raw = await file.read()
        data = json.loads(raw)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Файл должен быть корректным JSON")
    if not isinstance(data, list):
        raise HTTPException(status_code=400, detail="JSON должен быть списком")
    with db_connection() as conn:
        with conn.cursor() as cur:
            added = 0
            try:
                for m in data:
                    if "full_name" not in m or "kinship_type" not in m:
                        continue
                    cur.execute(
                        "INSERT INTO family_members (full_name, kinship_type) VALUES (%s, %s) RETURNING id",
                        (m["full_name"], m["kinship_type"]),
                    )
                    new_id = cur.fetchone()[0]
                    _add_default_income_types_for_member(cur, new_id)
                    added += 1
                conn.commit()
                log_audit({"role": "admin"}, "upload_members", details=f"Added {added} members")
            except Exception as e:
                logger.error(f"Bulk upload error: {e}")
                raise HTTPException(status_code=400, detail=str(e))
    return {"added": added}

@app.get("/members")
def list_members(token: str = Header(...)):
    user = session_manager.get_user(token)
    if user["role"] not in ("admin", "client"):
        raise HTTPException(status_code=403, detail="Доступ запрещён")
    with db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, full_name, kinship_type FROM family_members ORDER BY id")
            rows = cur.fetchall()
    return rows

@app.put("/members/{member_id}")
def update_member(member_id: int, data: MemberCreate, token: str = Header(...)):
    auth_proxy.check(token, "admin")
    with db_connection() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    "UPDATE family_members SET full_name = %s, kinship_type = %s WHERE id = %s RETURNING id",
                    (data.full_name, data.kinship_type.value, member_id),
                )
                if cur.rowcount == 0:
                    raise HTTPException(status_code=404, detail="Член семьи не найден")
                conn.commit()
                log_audit({"role": "admin"}, "update_member", entity_id=member_id,
                          details=f"Name: {data.full_name}, kinship: {data.kinship_type.value}")
            except Exception as e:
                raise HTTPException(status_code=400, detail=str(e))
    return {"detail": "Член семьи обновлён"}

@app.delete("/members/{member_id}")
def delete_member(member_id: int, token: str = Header(...)):
    auth_proxy.check(token, "admin")
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM income WHERE member_id = %s UNION SELECT id FROM expense WHERE operator_member_id = %s", (member_id, member_id))
            if cur.fetchone():
                raise HTTPException(status_code=400, detail="Нельзя удалить члена семьи, у которого есть доходы или расходы")
            cur.execute("DELETE FROM family_members WHERE id = %s RETURNING id", (member_id,))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Член семьи не найден")
            conn.commit()
            log_audit({"role": "admin"}, "delete_member", entity_id=member_id)
    return {"detail": "Член семьи удалён"}

# ---------- Income CRUD ----------
@app.post("/income")
def add_income(data: IncomeIn, token: str = Header(...)):
    user = session_manager.get_user(token)
    if user["role"] not in ("admin", "client"):
        raise HTTPException(status_code=403, detail="Требуется роль admin или client")
    if user["role"] == "admin":
        if data.member_id is None:
            raise HTTPException(status_code=400, detail="Для администратора необходимо указать member_id")
        member_id = data.member_id
    else:
        if data.member_id is not None:
            raise HTTPException(status_code=403, detail="Клиент не может указывать member_id для другого члена семьи")
        member_id = user["member_id"]
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM family_members WHERE id = %s", (member_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Член семьи не найден")
            cur.execute(
                "SELECT 1 FROM member_income_types WHERE member_id = %s AND income_type = %s",
                (member_id, data.income_type.value),
            )
            if not cur.fetchone():
                raise HTTPException(status_code=403, detail=f"Члену семьи не разрешён тип дохода '{data.income_type.value}'")
            try:
                income_type_id = _get_income_type_id(cur, data.income_type.value)
                account_id = _get_account_id(cur, data.account.value)
                operator_member_id = member_id if user["role"] == "admin" else user["member_id"]
                cur.execute(
                    """INSERT INTO income (member_id, income_type_id, account_id, amount, date, operator_member_id)
                       VALUES (%s, %s, %s, %s, %s, %s) RETURNING id""",
                    (member_id, income_type_id, account_id, data.amount, data.date, operator_member_id),
                )
                new_id = cur.fetchone()[0]
                conn.commit()
                log_audit(user, "add_income", entity_id=new_id,
                          details=f"Amount: {data.amount}, type: {data.income_type.value}, for member_id: {member_id}")
            except Exception as e:
                logger.error(f"Error adding income: {e}")
                raise HTTPException(status_code=400, detail=str(e))
    return {"detail": "Доход добавлен", "id": new_id}

@app.put("/income/{income_id}")
def update_income(income_id: int, data: IncomeIn, token: str = Header(...)):
    user = session_manager.get_user(token)
    if user["role"] not in ("admin", "client"):
        raise HTTPException(status_code=403, detail="Требуется роль admin или client")
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT member_id, operator_member_id FROM income WHERE id = %s", (income_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Доход не найден")
            if user["role"] != "admin" and row[0] != user["member_id"]:
                raise HTTPException(status_code=403, detail="Можно редактировать только свои доходы")
            income_type_id = _get_income_type_id(cur, data.income_type.value)
            account_id = _get_account_id(cur, data.account.value)
            op_member_id = row[1] if user["role"] == "admin" else user["member_id"]
            cur.execute(
                """UPDATE income 
                   SET amount = %s, date = %s, income_type_id = %s,
                       account_id = %s, operator_member_id = %s
                   WHERE id = %s""",
                (data.amount, data.date, income_type_id, account_id, op_member_id, income_id),
            )
            conn.commit()
            log_audit(user, "update_income", entity_id=income_id,
                      details=f"New amount: {data.amount}, type: {data.income_type.value}")
    return {"detail": "Доход обновлён"}

@app.delete("/income/{income_id}")
def delete_income(income_id: int, token: str = Header(...)):
    user = session_manager.get_user(token)
    if user["role"] not in ("admin", "client"):
        raise HTTPException(status_code=403, detail="Требуется роль admin или client")
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT member_id FROM income WHERE id = %s", (income_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Доход не найден")
            if user["role"] != "admin" and row[0] != user["member_id"]:
                raise HTTPException(status_code=403, detail="Можно удалять только свои доходы")
            cur.execute("DELETE FROM income WHERE id = %s", (income_id,))
            conn.commit()
            log_audit(user, "delete_income", entity_id=income_id)
    return {"detail": "Доход удалён"}

# ---------- Expense CRUD ----------
@app.post("/expense")
def add_expense(data: ExpenseIn, token: str = Header(...)):
    user = session_manager.get_user(token)
    if user["role"] not in ("admin", "client"):
        raise HTTPException(status_code=403, detail="Требуется роль admin или client")
    if user["role"] == "admin":
        if data.member_id is None:
            raise HTTPException(status_code=400, detail="Для администратора необходимо указать member_id")
        member_id = data.member_id
    else:
        if data.member_id is not None:
            raise HTTPException(status_code=403, detail="Клиент не может указывать member_id для другого члена семьи")
        member_id = user["member_id"]
    with db_connection() as conn:
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM family_members WHERE id = %s", (member_id,))
                if not cur.fetchone():
                    raise HTTPException(status_code=404, detail="Член семьи не найден")
                category_id = _get_category_id(cur, data.category.value)
                operator_member_id = member_id if user["role"] == "admin" else user["member_id"]
                cur.execute(
                    """INSERT INTO expense (category_id, amount, date, is_planned, operator_member_id)
                       VALUES (%s, %s, %s, %s, %s) RETURNING id""",
                    (category_id, data.amount, data.date, data.is_planned, operator_member_id),
                )
                new_id = cur.fetchone()[0]
                conn.commit()
                log_audit(user, "add_expense", entity_id=new_id,
                          details=f"Amount: {data.amount}, category: {data.category.value}, planned: {data.is_planned}, for member_id: {member_id}")
        except psycopg2.Error as e:
            conn.rollback()
            logger.error(f"Error adding expense: {e}")
            if "Запрет расхода" in str(e) or "balance" in str(e).lower():
                raise HTTPException(status_code=400, detail="Расход запрещён: текущий баланс <= 0")
            raise HTTPException(status_code=400, detail=f"Ошибка БД: {e}")
        except Exception as e:
            conn.rollback()
            logger.error(f"Unexpected error: {e}")
            raise HTTPException(status_code=400, detail=str(e))
    return {"detail": "Расход добавлен", "id": new_id}

@app.put("/expense/{expense_id}")
def update_expense(expense_id: int, data: ExpenseIn, token: str = Header(...)):
    user = session_manager.get_user(token)
    if user["role"] not in ("admin", "client"):
        raise HTTPException(status_code=403, detail="Требуется роль admin или client")
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT operator_member_id FROM expense WHERE id = %s", (expense_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Расход не найден")
            if user["role"] != "admin" and row[0] != user["member_id"]:
                raise HTTPException(status_code=403, detail="Можно редактировать только свои расходы")
            category_id = _get_category_id(cur, data.category.value)
            op_member_id = row[0] if user["role"] == "admin" else user["member_id"]
            cur.execute(
                """UPDATE expense 
                   SET amount = %s, date = %s, is_planned = %s,
                       category_id = %s, operator_member_id = %s
                   WHERE id = %s""",
                (data.amount, data.date, data.is_planned, category_id, op_member_id, expense_id),
            )
            conn.commit()
            log_audit(user, "update_expense", entity_id=expense_id,
                      details=f"New amount: {data.amount}, category: {data.category.value}")
    return {"detail": "Расход обновлён"}

@app.delete("/expense/{expense_id}")
def delete_expense(expense_id: int, token: str = Header(...)):
    user = session_manager.get_user(token)
    if user["role"] not in ("admin", "client"):
        raise HTTPException(status_code=403, detail="Требуется роль admin или client")
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT operator_member_id FROM expense WHERE id = %s", (expense_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Расход не найден")
            if user["role"] != "admin" and row[0] != user["member_id"]:
                raise HTTPException(status_code=403, detail="Можно удалять только свои расходы")
            cur.execute("DELETE FROM expense WHERE id = %s", (expense_id,))
            conn.commit()
            log_audit(user, "delete_expense", entity_id=expense_id)
    return {"detail": "Расход удалён"}

# ---------- Reporting ----------
def _get_incomes(start_date: Optional[date] = None, end_date: Optional[date] = None) -> list:
    with db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            query = """
                SELECT i.id, it.name AS income_type, i.amount, i.date,
                       a.name AS account, m.full_name, m.kinship_type
                FROM income i
                JOIN family_members m ON i.member_id = m.id
                JOIN income_types it ON i.income_type_id = it.id
                JOIN accounts a ON i.account_id = a.id
                WHERE 1=1
            """
            params = []
            if start_date:
                query += " AND i.date >= %s"
                params.append(start_date)
            if end_date:
                query += " AND i.date <= %s"
                params.append(end_date)
            query += " ORDER BY i.date, i.id"
            cur.execute(query, params)
            rows = cur.fetchall()
    return rows

def _get_expenses(start_date: Optional[date] = None, end_date: Optional[date] = None,
                  is_planned: Optional[bool] = None) -> list:
    with db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            query = """
                SELECT e.id, ec.name AS category, e.amount, e.date, e.is_planned
                FROM expense e
                JOIN expense_categories ec ON e.category_id = ec.id
                WHERE 1=1
            """
            params = []
            if start_date:
                query += " AND e.date >= %s"
                params.append(start_date)
            if end_date:
                query += " AND e.date <= %s"
                params.append(end_date)
            if is_planned is not None:
                query += " AND e.is_planned = %s"
                params.append(is_planned)
            query += " ORDER BY e.date, e.id"
            cur.execute(query, params)
            rows = cur.fetchall()
    return rows

@app.get("/incomes")
def list_incomes(start_date: date = None, end_date: date = None, token: str = Header(...)):
    session_manager.get_user(token)
    return _get_incomes(start_date, end_date)

@app.get("/expenses")
def list_expenses(start_date: date = None, end_date: date = None, is_planned: bool = None, token: str = Header(...)):
    session_manager.get_user(token)
    return _get_expenses(start_date, end_date, is_planned)

@app.get("/balance")
def get_balance(token: str = Header(...)):
    session_manager.get_user(token)
    with db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COALESCE(SUM(amount),0) FROM income")
            total_inc = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(SUM(amount),0) FROM expense WHERE is_planned = FALSE")
            total_exp = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(SUM(amount),0) FROM expense WHERE is_planned = TRUE")
            planned = cur.fetchone()[0]
    return {
        "total_income": float(total_inc),
        "total_expenses": float(total_exp),
        "planned_expenses": float(planned),
        "balance": float(total_inc) - float(total_exp),
    }

@app.get("/planned")
def get_planned_expenses(start_date: date = None, end_date: date = None, token: str = Header(...)):
    session_manager.get_user(token)
    return _get_expenses(start_date, end_date, is_planned=True)

@app.get("/report")
def get_report(start_date: date, end_date: date, token: str = Header(...)):
    session_manager.get_user(token)
    incomes = _get_incomes(start_date, end_date)
    expenses = _get_expenses(start_date, end_date, is_planned=False)
    total_income = sum(i["amount"] for i in incomes)
    total_expense = sum(e["amount"] for e in expenses)
    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "incomes": incomes,
        "expenses": expenses,
        "total_income": float(total_income),
        "total_expense": float(total_expense),
        "balance": float(total_income) - float(total_expense),
    }

# ---------- Excel Export ----------
def _format_amount_excel(amount: float) -> str:
    if amount == int(amount):
        return f"{int(amount):,}".replace(",", " ")
    return f"{amount:,.2f}".replace(",", " ")

class ExcelReportBuilder:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Семейный бюджет"

    def set_title(self, start_date: date, end_date: date):
        self.ws.merge_cells('A1:D1')
        title_cell = self.ws['A1']
        title_cell.value = f"Баланс семьи за период с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}:"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        self.ws.row_dimensions[1].height = 25

    def add_headers(self):
        headers = ["Статья", "Кто", "Сумма", "Дата"]
        self.ws.append(headers)
        header_font = Font(bold=True)
        header_border = Border(bottom=Side(style='thin'))
        for col_idx in range(1, 5):
            cell = self.ws.cell(row=2, column=col_idx)
            cell.font = header_font
            cell.border = header_border

    def add_income_section(self, incomes):
        self.ws.append(["Доход", "", "", ""])
        row = self.ws.max_row
        for col in range(1, 5):
            self.ws.cell(row=row, column=col).font = Font(bold=True)
        for inc in incomes:
            kto = f"{inc['full_name']} ({inc['kinship_type']})" if inc.get('full_name') else ""
            self.ws.append([
                inc['income_type'],
                kto,
                _format_amount_excel(float(inc['amount'])),
                inc['date'].strftime('%d.%m.%Y') if inc['date'] else ""
            ])

    def add_expense_section(self, expenses):
        self.ws.append(["Расход", "", "", ""])
        row = self.ws.max_row
        for col in range(1, 5):
            self.ws.cell(row=row, column=col).font = Font(bold=True)
        for exp in expenses:
            self.ws.append([
                exp['category'],
                "",
                _format_amount_excel(float(exp['amount'])),
                exp['date'].strftime('%d.%m.%Y') if exp['date'] else ""
            ])

    def add_totals(self, total_income, total_expense):
        self.ws.append(["Итого", "", "", ""])
        row_idx = self.ws.max_row
        for col in range(1, 5):
            self.ws.cell(row=row_idx, column=col).border = Border(top=Side(style='thin'))
            self.ws.cell(row=row_idx, column=col).font = Font(bold=True)

        self.ws.append(["Доход", "", _format_amount_excel(total_income), ""])
        self.ws.append(["Расход", "", _format_amount_excel(total_expense), ""])
        self.ws.append(["Баланс", "", _format_amount_excel(total_income - total_expense), ""])

        start_total_row = self.ws.max_row - 2
        for r in range(start_total_row, self.ws.max_row + 1):
            for col in range(1, 5):
                self.ws.cell(row=r, column=col).font = Font(bold=True)

    def finalise(self):
        self.ws.freeze_panes = 'A3'
        min_widths = {'A': 20, 'B': 30, 'C': 18, 'D': 15}
        for col_letter in ('A', 'B', 'C', 'D'):
            max_len = min_widths[col_letter]
            for cell in self.ws[col_letter]:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            self.ws.column_dimensions[col_letter].width = max_len + 2

    def build(self):
        self.finalise()
        stream = io.BytesIO()
        self.wb.save(stream)
        stream.seek(0)
        return stream

@app.get("/export/excel")
def export_excel(start_date: date, end_date: date, token: str = Header(...)):
    session_manager.get_user(token)
    try:
        incomes = _get_incomes(start_date, end_date)
        expenses = _get_expenses(start_date, end_date, is_planned=False)
        builder = ExcelReportBuilder()
        builder.set_title(start_date, end_date)
        builder.add_headers()
        builder.add_income_section(incomes)
        builder.add_expense_section(expenses)
        total_income = sum(i['amount'] for i in incomes)
        total_expense = sum(e['amount'] for e in expenses)
        builder.add_totals(total_income, total_expense)
        stream = builder.build()
        filename = f"Семейный_бюджет_{start_date.strftime('%d.%m.%Y')}_{end_date.strftime('%d.%m.%Y')}.xlsx"
        encoded_filename = quote(filename)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    except Exception as e:
        logger.error(f"Excel export error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ошибка при создании отчёта")

# ---------- ЭНДПОИНТЫ ДЛЯ ИСТОРИИ ----------
@app.get("/audit")
def get_audit(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    action: Optional[str] = None,
    token: str = Header(...)
):
    """Возвращает журнал аудита. Для admin – все записи, для client – только свои."""
    user = session_manager.get_user(token)
    if user["role"] not in ("admin", "client"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    
    with db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            query = """
                SELECT id, user_id, user_role, action, details, entity_id, created_at
                FROM audit_log
                WHERE 1=1
            """
            params = []
            if user["role"] != "admin":
                query += " AND user_id = %s"
                params.append(user["member_id"])
            if start_date:
                query += " AND created_at >= %s"
                params.append(start_date)
            if end_date:
                query += " AND created_at <= %s"
                params.append(end_date + timedelta(days=1))
            if action:
                query += " AND action = %s"
                params.append(action)
            query += " ORDER BY created_at DESC"
            cur.execute(query, params)
            rows = cur.fetchall()
    for row in rows:
        row["created_at"] = row["created_at"].isoformat()
    return rows

@app.get("/balance/history")
def get_balance_history(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    token: str = Header(...)
):
    """Возвращает историю изменения баланса. Для client – без operation_id."""
    user = session_manager.get_user(token)
    if user["role"] not in ("admin", "client"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    
    with db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if user["role"] == "admin":
                query = """
                    SELECT changed_at, operation_type, operation_id,
                           balance_before, balance_after, change_amount
                    FROM balance_history
                    WHERE 1=1
                """
            else:
                query = """
                    SELECT changed_at, operation_type,
                           balance_before, balance_after, change_amount
                    FROM balance_history
                    WHERE 1=1
                """
            params = []
            if start_date:
                query += " AND changed_at >= %s"
                params.append(start_date)
            if end_date:
                query += " AND changed_at <= %s"
                params.append(end_date + timedelta(days=1))
            query += " ORDER BY changed_at"
            cur.execute(query, params)
            rows = cur.fetchall()
    for row in rows:
        row["changed_at"] = row["changed_at"].isoformat()
    return rows

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("backend:app", host="127.0.0.1", port=8000, reload=False)